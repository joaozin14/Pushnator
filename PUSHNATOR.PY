# Importando bibliotecas
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import openpyxl
from openpyxl import Workbook
from time import sleep
from datetime import datetime

# Lendo os dados da planilha
workbook = openpyxl.load_workbook('Inserir_Push.xlsx') 
ler_planilha = workbook['Plan1']

total_processos = len([linha for linha in ler_planilha.iter_rows(min_row=1) if linha[0].value])
print(f"Total de processos na planilha: {total_processos}")
sleep(5)

# Criando uma nova planilha para armazenar os resultados
nova_planilha = Workbook()
planilha_sheet = nova_planilha.active
planilha_sheet.title = "Resultados"
planilha_sheet.append(["Processo", "Status"])

# Configurações do navegador
options = webdriver.ChromeOptions()
options.add_argument('--start-maximized')  # Iniciar maximizado
options.add_argument('--disable-infobars')  # Desabilitar infobars
options.add_argument('--disable-extensions')  # Desabilitar extensões
options.page_load_strategy = 'normal' # Vai carregar a página normalmente

# Inicializa os serviços do Chrome automaticamente usando as definições acima
driver = webdriver.Chrome(options=options)

# Abrindo PJE e fazendo o login
driver.get("https://pje1g.trf3.jus.br/pje/login.seam")

# A página de login tem um inframe que barra a automação, esse trecho burla essa inframe
frame_correta = driver.find_element(By.XPATH, '//*[@id="ssoFrame"]')
driver.switch_to.frame(frame_correta)

# Identificando os IDs e inserindo o login
login = driver.find_element(By.ID, 'username').send_keys('***********') # ALTERAR CPF AQUI
senha = driver.find_element(By.ID, 'password').send_keys('*********') # ALTERAR SENHA AQUI
sleep(2)
clic = driver.find_element(By.ID, 'kc-login').click()

# Esperar até que um elemento específico da nova página esteja presente, isso é uma forma de garantir que o login foi feito com sucesso
try:
   
    WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="barraSuperiorPrincipal"]/div[1]/div[1]/ul/li/a')))
    print("Login bem-sucedido e redirecionado para a nova página.")
except:
    print(f"Ocorreu um erro durante o login ou o redirecionamento")

# Redirecionar para uma URL específica após o login
nova_url = "https://pje1g.trf3.jus.br/pje/Push/listView.seam"
driver.get(nova_url)

#Lendo os dados da planilha
for indice, linha in enumerate(ler_planilha.iter_rows(min_row=1), start=1):  # enumerando a partir de 1
    processo = linha[0].value

#Encontrando o seletor e incluindo o processo nele    
    if processo:
        print(f"{indice}/{total_processos} processos")
        print(f"Analisando o processo {indice}: {processo}")
        print()
        
        seletor_processo = driver.find_element(By. XPATH, '//*[@id="j_id148:inputNumeroProcesso-inputNumeroProcessoDecoration:inputNumeroProcesso-inputNumeroProcesso"]')
        seletor_processo.clear()
        seletor_processo.send_keys(processo)
        print(f'Analisando o processo {processo}... ')
        sleep(3)

        botao_incluir = driver.find_element(By.XPATH, '//*[@id="j_id148:btnIncluirAcompanhamento"]')
        botao_incluir.click()
    
#Experando aparecer a caixa de mensagem com a informação sobre o processo
        try:
         WebDriverWait(driver, 360).until(EC.visibility_of_element_located((By.ID, "dialogMessage")))
        except:
         print(f"Não encontrado a caixa de mensagem")

#Pegando essa mensagem e a tranformando em texto para incluir na planilha
         msg = ''
        try:
            msg = driver.find_element(By.CSS_SELECTOR, 'span.rich-messages-label').text
            sleep(3)
        except:
            print('Não encontrado a mensagem')
    
    
        print(f"{processo} - {msg}\n")
#Atribuindo a planilha o processo e sua mensagem    
        planilha_sheet.append([processo, msg])

driver.quit()

# Jogando as informações na planilha nova e a nomeando
data_atual = datetime.now().strftime('%d-%m-%Y')
nome_arquivo = f'Resultados-Inclusao-{data_atual}.xlsx'
nova_planilha.save(nome_arquivo)
